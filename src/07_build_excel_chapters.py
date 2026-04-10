import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Directory layout ─────────────────────────────────────────────────────────
import pathlib as _pl
CSV_DIR  = _pl.Path('csv')    # input CSVs:  csv/books_metadata_full.csv, csv/books_text_*.csv
JSON_DIR = _pl.Path('json')   # all JSON/JSONL files
JSON_DIR.mkdir(exist_ok=True)


with open(str(JSON_DIR / 'nlp_results_chapters.json')) as f:  R = json.load(f)
try:
    with open(str(JSON_DIR / 'nlp_results.json')) as _f: _RB = json.load(_f)
    R['topic_names'] = _RB.get('topic_names')  # carry book-level names
except Exception: pass
with open(str(JSON_DIR / 'summaries.json')) as f:  S = json.load(f)

n_topics   = R['n_topics']
n_clusters = R['best_k']
book_ids   = R['book_ids']
chs_meta   = R['chapters']
dom_topics = R['dominant_topics']
clusters   = R['cluster_labels']
keyphrases = R['keyphrases']
top_words  = R['top_words']
book_tc    = R['book_topic_counts']

_BASE_TOPIC_NAMES = [
    'Human & Social Experience',
    'Mathematical & Formal Systems',
    'General Systems Theory',
    'History & Philosophy of Cybernetics',
    'Management & Organisational Cybernetics',
    'Control Theory & Engineering',
    'Topic 7', 'Topic 8', 'Topic 9',
]
_carried = R.get('topic_names') or _BASE_TOPIC_NAMES
TOPIC_NAMES = (_carried + [f'Topic {i+1}' for i in range(len(_carried), n_topics)])[:n_topics]

FILLS = [
    PatternFill('solid',fgColor='DBEAFE'), PatternFill('solid',fgColor='DCFCE7'),
    PatternFill('solid',fgColor='FEF3C7'), PatternFill('solid',fgColor='EDE9FE'),
    PatternFill('solid',fgColor='FEE2E2'), PatternFill('solid',fgColor='CFFAFE'),
    PatternFill('solid',fgColor='FCE7F3'), PatternFill('solid',fgColor='F0FDF4'),
    PatternFill('solid',fgColor='FFF7ED'), PatternFill('solid',fgColor='E0F2FE'),
    PatternFill('solid',fgColor='F0FDF4'), PatternFill('solid',fgColor='FEF9C3'),
]
HDR = PatternFill('solid',fgColor='1E3A5F')
HDR_FONT = Font(name='Calibri',bold=True,color='FFFFFF',size=10)
TITLE_FONT = Font(name='Calibri',bold=True,size=11)
NORM_FONT  = Font(name='Calibri',size=9)
WRAP = Alignment(wrap_text=True,vertical='top')
thin = Side(style='thin',color='D1D5DB')
BORDER = Border(left=thin,right=thin,top=thin,bottom=thin)

wb = openpyxl.Workbook()

# ── Sheet 1: Chapter Results ──────────────────────────────────────────────────
ws1 = wb.active; ws1.title = 'Chapter Results'
hdrs = ['Chapter ID','Book','Author','Ch #','Chapter Title','Word Count',
        'Dominant Topic','Topic Name','Cluster','Top Keyphrases','Summary']
ws1.row_dimensions[1].height = 22
for c, h in enumerate(hdrs, 1):
    cell = ws1.cell(1, c, h)
    cell.font = HDR_FONT; cell.fill = HDR; cell.border = BORDER
    cell.alignment = Alignment(horizontal='center',vertical='center')

for i, ch in enumerate(chs_meta):
    t   = dom_topics[i]
    cl  = clusters[i]
    cid = ch['chapter_id']
    kps = ', '.join(keyphrases.get(cid,[])[:6])
    summ = ''
    for s_ch in S.get(ch['book_id'],{}).get('chapters',[]):
        if s_ch['index'] == ch['ch_index']:
            summ = s_ch.get('summary','')
            break
    row = [cid, ch['book_title'], ch['book_author'], ch['ch_index'],
           ch['ch_title'], ch['word_count'], f"T{t+1}", TOPIC_NAMES[t],
           f"C{cl+1}", kps, summ]
    ws1.row_dimensions[i+2].height = 40
    for c, v in enumerate(row, 1):
        cell = ws1.cell(i+2, c, v)
        cell.font = NORM_FONT; cell.fill = FILLS[t % len(FILLS)]
        cell.border = BORDER; cell.alignment = WRAP

widths = [18,35,18,6,40,8,10,24,8,45,80]
for c, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

ws1.freeze_panes = 'A2'

# ── Sheet 2: NMF Topics ───────────────────────────────────────────────────────
ws2 = wb.create_sheet('NMF Topics')
h2 = ['Topic #','Topic Name','Top Keywords','Chapter Count']
for c, h in enumerate(h2, 1):
    cell = ws2.cell(1,c,h); cell.font=HDR_FONT; cell.fill=HDR; cell.border=BORDER
for t in range(n_topics):
    cnt = sum(1 for d in dom_topics if d==t)
    row = [f"T{t+1}", TOPIC_NAMES[t], ', '.join(top_words[t][:12]), cnt]
    for c, v in enumerate(row, 1):
        cell = ws2.cell(t+2,c,v)
        cell.font=NORM_FONT; cell.fill=FILLS[t]; cell.border=BORDER
for c, w in enumerate([8,28,70,12],1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ── Sheet 3: Book × Topic ─────────────────────────────────────────────────────
ws3 = wb.create_sheet('Book × Topic')
ws3.cell(1,1,'Book')
for t in range(n_topics):
    cell = ws3.cell(1,t+2,f"T{t+1} {TOPIC_NAMES[t]}")
    cell.font=HDR_FONT; cell.fill=HDR; cell.border=BORDER
ws3.cell(1,1,'Book').font=HDR_FONT; ws3.cell(1,1).fill=HDR; ws3.cell(1,1).border=BORDER
for i, bid in enumerate(book_ids):
    ws3.cell(i+2,1,S[bid]['title'][:50]).font=NORM_FONT
    ws3.cell(i+2,1).border=BORDER
    tc = book_tc.get(bid,[0]*n_topics)
    for t in range(n_topics):
        cell = ws3.cell(i+2,t+2,tc[t])
        cell.font=NORM_FONT; cell.fill=FILLS[t]; cell.border=BORDER
        cell.alignment=Alignment(horizontal='center')
ws3.column_dimensions['A'].width = 52
for t in range(n_topics):
    ws3.column_dimensions[get_column_letter(t+2)].width = 22

# ── Sheet 4: Clusters ─────────────────────────────────────────────────────────
ws4 = wb.create_sheet('Clusters')
h4 = ['Cluster','Chapter Count','Dominant Topic','Books in Cluster']
for c, h in enumerate(h4,1):
    cell = ws4.cell(1,c,h); cell.font=HDR_FONT; cell.fill=HDR; cell.border=BORDER
for cl in range(n_clusters):
    cids = [i for i, c in enumerate(clusters) if c==cl]
    books_in = {}
    for i in cids:
        bid = chs_meta[i]['book_id']
        books_in[bid] = books_in.get(bid,0)+1
    top_t = max(set(dom_topics[i] for i in cids), key=lambda t: sum(1 for i in cids if dom_topics[i]==t))
    bk_str = '; '.join(f"{S[bid]['title'][:25]}({books_in[bid]})" for bid in sorted(books_in, key=books_in.get, reverse=True)[:6])
    row = [f"C{cl+1}", len(cids), f"T{top_t+1} {TOPIC_NAMES[top_t]}", bk_str]
    for c, v in enumerate(row,1):
        cell = ws4.cell(cl+2,c,v); cell.font=NORM_FONT; cell.border=BORDER
        cell.alignment=WRAP; cell.fill=FILLS[top_t%len(FILLS)]
for c, w in enumerate([10,12,28,90],1):
    ws4.column_dimensions[get_column_letter(c)].width = w

out = 'data/outputs/book_nlp_chapters.xlsx'
wb.save(out)
print(f"Excel saved: {out}")